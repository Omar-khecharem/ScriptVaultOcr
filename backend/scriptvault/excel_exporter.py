"""Export de lots de résultats OCR en classeur Excel (.xlsx) — projet SWIT.

Génération **100 % hors-ligne** (aucun appel réseau) d'un classeur multi-
onglets :

* Onglet **« Données Extraites »** — un tableau avec une ligne par page
  analysée (TIF multi-pages, PDF, images) : entêtes stylisés (fond bleu
  marine ``#1F4E78``, texte blanc gras, centré), filtres activés, gel du
  bandeau d'entête, ajustement automatique des largeurs de colonnes et mise
  en forme conditionnelle appliquée à l'écriture (confiance, statut,
  code-barres).
* Onglet **« Rapport de Traitement »** — synthèse des indicateurs clés :
  nombre total de fichiers traités, taux d'acceptation automatique (%),
  nombre de documents nécessitant révision, confiance moyenne, détection de
  codes-barres.

Performances : ``openpyxl`` en **mode écriture rapide**
(``Workbook(write_only=True)``) avec ``WriteOnlyCell`` stylisées — les styles
sont posés une fois par valeur pendant l'écriture, sans re-parcours du
classeur : des volumes massifs de TIF/PDF peuvent être exportés localement.

Le module expose un modèle de données minimal (dataclasses) déconnecté de
FastAPI, réutilisable tel quel par le desktop ou un CLI.

Exemple::

    from excel_exporter import ExcelDocument, ExcelPage, ExcelField, export_excel

    docs = [
        ExcelDocument(
            filename="scan_2024.tif",
            pages=[
                ExcelPage(
                    page=1,
                    confidence=0.96,
                    fields=[
                        ExcelField(text="DUPONT", label="nom"),
                        ExcelField(text="AB123456", label="cin"),
                    ],
                    barcodes=[ExcelBarcode(data="25052024", type="EAN-13")],
                )
            ],
        )
    ]
    with open("export.xlsx", "wb") as handle:
        handle.write(export_excel(docs))
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional


# --------------------------------------------------------------------------- #
# Modèle de données (indépendant de l'API)
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class ExcelField:
    """Une zone OCR étiquetée (Nom, Prénom, CIN, Identifiant, ...)."""

    text: str = ""
    confidence: float = 0.0
    label: Optional[str] = None


@dataclass(frozen=True)
class ExcelBarcode:
    """Un code-barres / QR détecté localement (fiabilité 100 % présumée)."""

    data: str
    type: str = "BARCODE"


@dataclass(frozen=True)
class ExcelPage:
    """Une page analysée — une ligne du tableau « Données Extraites »."""

    page: int = 1
    text: str = ""
    confidence: float = 0.0
    fields: list[ExcelField] = field(default_factory=list)
    barcodes: list[ExcelBarcode] = field(default_factory=list)


@dataclass(frozen=True)
class ExcelDocument:
    """Un document source (fichier TIF / PDF / image) et ses pages."""

    filename: str
    pages: list[ExcelPage] = field(default_factory=list)


class ExcelError(RuntimeError):
    """Échec de génération du classeur Excel."""


# --------------------------------------------------------------------------- #
# Règles visuelles & métier (constantes centralisées, réglables ici)
# --------------------------------------------------------------------------- #
_HEADER_BG = "1F4E78"  # bleu marine (entêtes)
_HEADER_FG = "FFFFFF"  # blanc (texte des entêtes)
_GOOD_FILL = "FFEB9C"  # orange : ligne validée automatiquement
_BAD_FILL = "FFC7CE"  # rouge : ligne nécessitant révision
_BARCODE_FILL = "C6EFCE"  # vert : données code-barres (fiabilité 100 %)
_KPI_FILL = "DDEBF7"  # bleu clair : valeurs du tableau de bord

# Seuil d'acceptation automatique (confiance globale d'une page).
ACCEPT_THRESHOLD = 0.85

# Règle métier (formulaire SWIT) : le CIN est alphanumérique, >= 4 caractères.
_CIN_PATTERN = re.compile(r"^[A-Za-z0-9]{4,}$")

# Ordre d'affichage des colonnes de l'onglet « Données Extraites ».
COLUMNS = [
    "Nom du Fichier",
    "N° Page",
    "Code-Barres / QR",
    "Nom",
    "Prénom",
    "CIN",
    "Identifiant",
    "Indice de Confiance Global (%)",
    "Statut Validation",
    "Date Traitement",
]

# Alias de libellés ROI → colonne canonique.
_LABEL_ALIASES = {
    "nom": "Nom",
    "name": "Nom",
    "lastname": "Nom",
    "prenom": "Prénom",
    "firstname": "Prénom",
    "cin": "CIN",
    "identifiant": "Identifiant",
    "id": "Identifiant",
    "identifier": "Identifiant",
    "no_dossier": "Identifiant",
}

_SAFE_FILENAME = re.compile(r"[^A-Za-z0-9._ -]")


# --------------------------------------------------------------------------- #
# Helpers de normalisation
# --------------------------------------------------------------------------- #
def _norm_label(label: str) -> Optional[str]:
    """Normalise un libellé ROI vers une colonne canonique (None si inconnu)."""
    return _LABEL_ALIASES.get(label.strip().lower())


def _field_text(fields: list[ExcelField], label: str) -> str:
    """Concatène les textes des zones portant ``label`` (alias tolérés)."""
    target = _norm_label(label)
    if target is None:
        return ""
    parts: list[str] = []
    for entry in fields:
        entry_label = (entry.label or "").strip()
        if entry_label and _norm_label(entry_label) == target:
            text = (entry.text or "").strip()
            if text and text not in parts:
                parts.append(text)
    return ", ".join(parts)


def _barcode_summary(barcodes: list[ExcelBarcode]) -> str:
    """Résumé lisible des codes-barres détectés sur une page."""
    items: list[str] = []
    for entry in barcodes:
        data = (entry.data or "").strip()
        if not data:
            continue
        items.append(f"{data} ({entry.type})" if entry.type else data)
    return " | ".join(items)


def _has_barcode(barcodes: list[ExcelBarcode]) -> bool:
    return bool(_barcode_summary(barcodes))


def _evaluate_status(
    confidence: float,
    fields: list[ExcelField],
    barcodes: list[ExcelBarcode],
    page_text: str,
) -> tuple[str, list[str]]:
    """Évalue la page et retourne ``(statut, violations)``.

    Règles d'acceptation automatique : confiance >= :data:`ACCEPT_THRESHOLD`
    **et** aucune règle métier violée (CIN mal formé, page sans données).
    Les codes-barres sont réputés fiables à 100 % mais ne lèvent pas seuls
    le seuil de confiance.
    """
    violations: list[str] = []
    if confidence < ACCEPT_THRESHOLD:
        violations.append(f"Confiance < {ACCEPT_THRESHOLD * 100:.0f} %")
    cin = _field_text(fields, "CIN")
    if cin and not _CIN_PATTERN.fullmatch(cin):
        violations.append("Format CIN invalide")
    has_payload = bool(
        _field_text(fields, "Nom")
        or _field_text(fields, "Prénom")
        or cin
        or _field_text(fields, "Identifiant")
        or (page_text or "").strip()
    )
    if not has_payload and not _has_barcode(barcodes):
        violations.append("Aucune donnée détectée sur la page")
    if violations:
        return "À réviser", violations
    return "Validé (Auto)", violations


def _display_width(text: Any, min_width: float = 8.0, max_width: float = 52.0) -> float:
    """Largeur d'affichage approximative d'une valeur (Unicode compris)."""
    if text is None:
        return min_width
    width = 0.0
    for ch in str(text):
        width += 1.0 if ord(ch) < 0x80 else 1.7
    return max(min_width, min(max_width, width + 2.0))


# --------------------------------------------------------------------------- #
# Fabrication du classeur
# --------------------------------------------------------------------------- #
def export_excel(
    documents: list[ExcelDocument],
    *,
    generated_at: Optional[datetime] = None,
) -> bytes:
    """Sérialise un lot de résultats OCR en classeur ``.xlsx``.

    Chaque page d'un document produit une ligne. Les zones étiquetées (ROI :
    Nom, Prénom, CIN, Identifiant) peuplent leurs colonnes ; les codes-barres
    / QR détectés sont mis en valeur (fiabilité 100 %).

    Args:
        documents: Lot de documents à exporter (jamais vide).
        generated_at: Date/heure de traitement (défaut : maintenant).

    Returns:
        Octets du fichier ``.xlsx`` prêts à être transmis.

    Raises:
        ExcelError: ``openpyxl`` absent, lot vide ou erreur de génération.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.cell.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ExcelError(
            "openpyxl est requis pour l'export Excel (pip install openpyxl)"
        ) from exc

    if not documents:
        raise ExcelError("Aucun document à exporter (lot vide).")
    generated = generated_at or datetime.now()

    # --- Styles construits une seule fois -------------------------------- #
    header_font = Font(bold=True, color=_HEADER_FG)
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")
    good_fill = PatternFill("solid", fgColor=_GOOD_FILL)
    bad_fill = PatternFill("solid", fgColor=_BAD_FILL)
    good_font = Font(color="9C6500", bold=True)
    bad_font = Font(color="9C0006", bold=True)
    barcode_fill = PatternFill("solid", fgColor=_BARCODE_FILL)
    barcode_font = Font(color="006100", bold=True)
    kpi_fill = PatternFill("solid", fgColor=_KPI_FILL)
    kpi_font = Font(bold=True)
    title_font = Font(bold=True, color=_HEADER_FG, size=13)
    title_fill = PatternFill("solid", fgColor=_HEADER_BG)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def cell(
        ws: Any,
        value: Any,
        *,
        font: Optional[Any] = None,
        fill: Optional[Any] = None,
        alignment: Optional[Any] = None,
        number_format: Optional[str] = None,
    ) -> Any:
        styled_cell = WriteOnlyCell(ws, value=value)
        if font is not None:
            styled_cell.font = font
        if fill is not None:
            styled_cell.fill = fill
        if alignment is not None:
            styled_cell.alignment = alignment
        if number_format is not None:
            styled_cell.number_format = number_format
        return styled_cell

    try:
        workbook = Workbook(write_only=True)

        # ================================================================ #
        # Onglet 1 — Données Extraites
        # ================================================================ #
        ws_data = workbook.create_sheet("Données Extraites")
        ws_data.freeze_panes = "A2"

        ws_data.append(
            [
                cell(
                    ws_data,
                    name,
                    font=header_font,
                    fill=header_fill,
                    alignment=header_align,
                )
                for name in COLUMNS
            ]
        )
        column_widths: list[float] = [_display_width(name) for name in COLUMNS]

        data_rows = 0
        for document in documents:
            pages = document.pages if document.pages else [ExcelPage(page=1)]
            for page in pages:
                fields = list(page.fields)
                barcodes = list(page.barcodes)
                confidence = max(0.0, min(1.0, float(page.confidence)))
                status, _violations = _evaluate_status(
                    confidence, fields, barcodes, page.text or ""
                )
                reliable = _has_barcode(barcodes)
                auto_ok = status == "Validé (Auto)"

                values: list[Any] = [
                    document.filename,
                    int(page.page) if page.page else 1,
                    _barcode_summary(barcodes),
                    _field_text(fields, "Nom"),
                    _field_text(fields, "Prénom"),
                    _field_text(fields, "CIN"),
                    _field_text(fields, "Identifiant"),
                    confidence,
                    status,
                    generated.strftime("%d/%m/%Y %H:%M"),
                ]
                rows: list[Any] = []
                for index, value in enumerate(values):
                    name = COLUMNS[index]
                    if name == "Nom du Fichier":
                        rows.append(cell(ws_data, value, alignment=left_wrap))
                    elif name in ("Nom", "Prénom", "CIN", "Identifiant"):
                        rows.append(cell(ws_data, value, alignment=left_wrap))
                    elif name == "Code-Barres / QR":
                        rows.append(cell(ws_data, value, alignment=center))
                    elif name == "Indice de Confiance Global (%)":
                        rows.append(
                            cell(
                                ws_data,
                                value,
                                alignment=center,
                                number_format="0.0%",
                                fill=good_fill if auto_ok else bad_fill,
                                font=good_font if auto_ok else bad_font,
                            )
                        )
                    elif name == "Statut Validation":
                        rows.append(
                            cell(
                                ws_data,
                                value,
                                alignment=center,
                                fill=good_fill if auto_ok else bad_fill,
                                font=good_font if auto_ok else bad_font,
                            )
                        )
                    else:
                        rows.append(cell(ws_data, value, alignment=center))

                # Les données issues d'un code-barres sont fiables à 100 % :
                # mise en valeur en vert gras.
                if reliable:
                    rows[2] = cell(
                        ws_data,
                        values[2],
                        alignment=center,
                        fill=barcode_fill,
                        font=barcode_font,
                    )
                ws_data.append(rows)
                data_rows += 1
                for col_index, value in enumerate(values):
                    column_widths[col_index] = max(
                        column_widths[col_index], _display_width(value)
                    )

        # Filtres automatiques sur toute la plage de données.
        last_col = get_column_letter(len(COLUMNS))
        ws_data.auto_filter.ref = f"A1:{last_col}{data_rows + 1}"
        for col_index, name in enumerate(COLUMNS):
            ws_data.column_dimensions[get_column_letter(col_index + 1)].width = min(
                column_widths[col_index] + 2.0, 60.0
            )

        # ================================================================ #
        # Onglet 2 — Rapport de Traitement (synthèse)
        # ================================================================ #
        ws_report = workbook.create_sheet("Rapport de Traitement")

        total_files = len(documents)
        total_pages = data_rows
        valid_rows = 0
        confidence_sum = 0.0
        barcode_files = 0
        for document in documents:
            for page in document.pages:
                page_confidence = max(0.0, min(1.0, float(page.confidence)))
                confidence_sum += page_confidence
                page_status, _ = _evaluate_status(
                    page_confidence,
                    list(page.fields),
                    list(page.barcodes),
                    page.text or "",
                )
                if page_status == "Validé (Auto)":
                    valid_rows += 1
            if any(_has_barcode(list(page.barcodes)) for page in document.pages):
                barcode_files += 1

        acceptance = (valid_rows / total_pages * 100.0) if total_pages else 0.0
        average = (confidence_sum / total_pages * 100.0) if total_pages else 0.0
        revision_count = total_pages - valid_rows

        # Titre du rapport (colonne A stylisée — les feuilles en mode
        # écriture rapide ne supportent pas la fusion de cellules).
        ws_report.append(
            [
                cell(
                    ws_report,
                    "Rapport de Traitement — ScriptVault OCR (SWIT)",
                    font=title_font,
                    fill=title_fill,
                ),
                cell(ws_report, None),
                cell(ws_report, None),
            ]
        )
        ws_report.append([cell(ws_report, None)])
        ws_report.append(
            [
                cell(
                    ws_report,
                    "Indicateur",
                    font=header_font,
                    fill=header_fill,
                    alignment=header_align,
                ),
                cell(
                    ws_report,
                    "Valeur",
                    font=header_font,
                    fill=header_fill,
                    alignment=header_align,
                ),
                cell(
                    ws_report,
                    "Remarque",
                    font=header_font,
                    fill=header_fill,
                    alignment=header_align,
                ),
            ]
        )
        ws_report.append(
            [
                cell(ws_report, None),
                cell(ws_report, None),
                cell(ws_report, None),
            ]
        )

        kpi_rows: list[tuple[str, Any, Optional[str]]] = [
            ("Nombre total de fichiers traités (TIF/PDF/Images)", total_files, None),
            ("Nombre total de pages analysées", total_pages, None),
            (
                "Taux d'acceptation automatique (%)",
                round(acceptance, 1),
                "Confiance >= 85 % et règles métier respectées",
            ),
            (
                "Nombre de documents nécessitant révision",
                revision_count,
                "Révision manuelle requise",
            ),
            ("Confiance moyenne globale (%)", round(average, 1), None),
            (
                "Fichiers avec code-barres détecté",
                barcode_files,
                "Données fiables à 100 %",
            ),
        ]
        for label, value, note in kpi_rows:
            ws_report.append(
                [
                    cell(ws_report, label, alignment=left_wrap),
                    cell(
                        ws_report, value, font=kpi_font, fill=kpi_fill, alignment=center
                    ),
                    cell(ws_report, note, alignment=left_wrap)
                    if note
                    else cell(ws_report, None),
                ]
            )

        ws_report.append([cell(ws_report, None)])
        ws_report.append(
            [
                cell(ws_report, "Date de génération du rapport", alignment=left_wrap),
                cell(
                    ws_report,
                    generated.strftime("%d/%m/%Y %H:%M"),
                    font=kpi_font,
                    alignment=center,
                ),
                cell(ws_report, None),
            ]
        )
        ws_report.column_dimensions["A"].width = 52.0
        ws_report.column_dimensions["B"].width = 16.0
        ws_report.column_dimensions["C"].width = 42.0

        # ------------------------------------------------------------------ #
        # Sérialisation en mémoire (aucune écriture disque).
        # ------------------------------------------------------------------ #
        import io

        buffer = io.BytesIO()
        try:
            workbook.save(buffer)
        finally:
            workbook.close()
        return buffer.getvalue()
    except ExcelError:
        raise
    except Exception as exc:
        raise ExcelError(f"Génération du classeur Excel échouée: {exc}") from exc
