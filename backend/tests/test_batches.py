"""Tests du workflow de traitement par lots (API /api/batches).

Le moteur PaddlePaddle n'est pas requis : une fabrique factice est injectée.
Les lots sont traités en tâche de fond — les tests sondent la progression
jusqu'à épuisement (le moteur factice est quasi instantané).
"""

from __future__ import annotations

import io
import time
from pathlib import Path

import cv2
import numpy as np
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from scriptvault.api.app import create_app
from scriptvault.config import Settings


def _fake_engine_factory():
    """Moteur factice : retourne deux lignes fixes, sans PaddlePaddle."""

    class FakeEngine:
        cpu_threads = 4
        is_ready = True

        def predict_array(self, image, *, preprocess=True):
            assert isinstance(image, np.ndarray)
            return [
                {
                    "text": "Nom : DUPONT",
                    "confidence": 0.99,
                    "box": [[0, 0], [50, 0], [50, 10], [0, 10]],
                },
                {
                    "text": "Prénom : Jean",
                    "confidence": 0.85,
                    "box": [[0, 20], [40, 20], [40, 30], [0, 30]],
                },
            ]

        predict_bytes = predict_array

        def close(self):
            pass

    return FakeEngine()


@pytest.fixture()
def settings(tmp_path: Path) -> Settings:
    return Settings(
        preload=False,
        max_concurrency=1,
        timeout_ms=5000,
        max_file_mb=5,
        preprocess=True,
        storage_root=str(tmp_path / "storage"),
    )


@pytest.fixture()
def client(settings: Settings) -> TestClient:
    app = create_app(settings=settings, engine_factory=_fake_engine_factory)
    with TestClient(app) as test_client:
        yield test_client


def _sample_image(tmp_path: Path, name: str = "scan.png") -> Path:
    image = np.full((300, 800, 3), 255, dtype=np.uint8)
    cv2.putText(image, "Texte", (40, 150), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (0, 0, 0), 3)
    path = tmp_path / name
    path.write_bytes(cv2.imencode(".png", image)[1].tobytes())
    return path


def _wait_done(client: TestClient, job_id: str, timeout: float = 8.0) -> dict:
    """Sonde le lot jusqu'à son terme (done / cancelled / error)."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        response = client.get(f"/api/batches/{job_id}")
        assert response.status_code == 200
        data = response.json()
        if data["status"] in ("done", "cancelled", "error"):
            return data
        time.sleep(0.05)
    raise AssertionError(f"Lot {job_id} toujours en cours après {timeout}s.")


def _create_batch(
    client: TestClient, images: list[Path], name: str = "Lot test"
) -> dict:
    files = [
        ("files", (image.name, image.read_bytes(), "image/png")) for image in images
    ]
    response = client.post("/api/batches", files=files, data={"name": name})
    assert response.status_code == 201, response.text
    return response.json()


# --------------------------------------------------------------------------- #
# Création & progression
# --------------------------------------------------------------------------- #
def test_create_batch_returns_job_summary(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    payload = _create_batch(client, [image])
    job = payload["job"]
    assert job["name"] == "Lot test"
    assert job["counts"]["total"] == 1
    assert job["status"] in ("pending", "processing", "done")

    summary = _wait_done(client, job["id"])
    assert summary["status"] == "done"
    assert summary["counts"]["done"] == 1
    assert summary["counts"]["error"] == 0
    assert summary["avg_confidence"] > 0.0


def test_batch_files_paginated_and_filtered(client: TestClient, tmp_path: Path):
    images = [_sample_image(tmp_path, f"scan_{i}.png") for i in range(3)]
    job = _create_batch(client, images)["job"]
    _wait_done(client, job["id"])

    response = client.get(
        f"/api/batches/{job['id']}/files", params={"page": 1, "page_size": 2}
    )
    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 3
    assert len(data["items"]) == 2
    assert all(item["status"] == "done" for item in data["items"])

    page_two = client.get(
        f"/api/batches/{job['id']}/files", params={"page": 2, "page_size": 2}
    ).json()
    assert len(page_two["items"]) == 1

    search = client.get(
        f"/api/batches/{job['id']}/files", params={"q": "scan_2"}
    ).json()
    assert search["total"] == 1
    assert search["items"][0]["name"] == "scan_2.png"


def test_file_detail_contains_pages_and_form(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]
    _wait_done(client, job["id"])

    listing = client.get(f"/api/batches/{job['id']}/files").json()
    file_id = listing["items"][0]["id"]

    detail = client.get(f"/api/batches/{job['id']}/files/{file_id}")
    assert detail.status_code == 200
    payload = detail.json()
    assert payload["status"] == "done"
    assert payload["pages"], "au moins une page OCR"
    page = payload["pages"][0]
    assert page["page"] == 1
    assert page["width"] > 0 and page["height"] > 0
    assert page["text"]
    assert "form" in page
    assert page["form"]["fields"] is not None


def test_file_preview_returns_png_data_url(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]
    _wait_done(client, job["id"])
    file_id = client.get(f"/api/batches/{job['id']}/files").json()["items"][0]["id"]

    response = client.get(f"/api/batches/{job['id']}/files/{file_id}/preview")
    assert response.status_code == 200
    preview = response.json()["preview"]
    assert preview.startswith("data:image/jpeg;base64,")

    missing = client.get(
        f"/api/batches/{job['id']}/files/{file_id}/preview", params={"page": 99}
    )
    assert missing.status_code == 404


def test_preview_persisted_once_on_disk(client: TestClient, tmp_path: Path):
    """Les aperçus JPEG sont écrits au moment de l'analyse et lus ensuite du
    disque : aucun re-prétraitement à chaque consultation."""
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]
    _wait_done(client, job["id"])

    file_id = client.get(f"/api/batches/{job['id']}/files").json()["items"][0]["id"]
    first = client.get(f"/api/batches/{job['id']}/files/{file_id}/preview")
    assert first.status_code == 200
    assert first.json()["preview"].startswith("data:image/jpeg;base64,")

    previews_dir = tmp_path / "storage" / "jobs" / job["id"] / "previews"
    assert previews_dir.is_dir()
    assert (previews_dir / "00001" / "0001.jpg").is_file(), (
        "l'aperçu doit être persisté sur disque (par fichier)"
    )


def test_previews_come_from_their_own_file(client: TestClient, tmp_path: Path):
    """Chaque fichier renvoie SON aperçu : le dossier d'aperçus est isolé par
    fichier (régression : auparavant tous les fichiers partageaient un dossier
    commun et renvoyaient l'image du premier fichier du lot)."""
    image_a = _sample_image(tmp_path, name="a.png")
    image_b = _sample_image(tmp_path, name="b.png")
    image_b.write_bytes(
        cv2.imencode(
            ".png",
            np.hstack([np.full((300, 400, 3), 0, dtype=np.uint8), np.full((300, 400, 3), 255, dtype=np.uint8)]),
        )[1].tobytes()
    )
    job = _create_batch(client, [image_a, image_b])["job"]
    _wait_done(client, job["id"])

    items = client.get(f"/api/batches/{job['id']}/files").json()["items"]
    assert len(items) == 2
    previews = [
        client.get(f"/api/batches/{job['id']}/files/{item['id']}/preview")
        .json()["preview"]
        for item in items
    ]
    assert previews[0] != previews[1], "les aperçus des fichiers doivent différer"

    previews_dir = tmp_path / "storage" / "jobs" / job["id"] / "previews"
    first_jpeg = previews_dir / "00001" / "0001.jpg"
    second_jpeg = previews_dir / "00002" / "0001.jpg"
    assert first_jpeg.is_file() and second_jpeg.is_file()
    assert first_jpeg.read_bytes() != second_jpeg.read_bytes()


def test_batch_export_excel(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]
    _wait_done(client, job["id"])

    response = client.get(f"/api/batches/{job['id']}/export.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert response.content.startswith(b"PK")


def test_excel_columns_are_filled_and_accent_safe(
    client: TestClient, tmp_path: Path
):
    """Régression : la colonne « Prénom » (accent) doit être peuplée, de même
    que les colonnes du gabarit (Série, Date de naissance, Établissement…),
    y compris via une correction manuelle d'un champ non lu par l'OCR."""
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]
    _wait_done(client, job["id"])
    file_id = client.get(f"/api/batches/{job['id']}/files").json()["items"][0]["id"]

    # L'OCR lit « Prénom Jean » : même sans correction, la colonne doit sortir.
    response = client.get(f"/api/batches/{job['id']}/export.xlsx")
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook["Données Extraites"]
    headers = [entry.value for entry in sheet[1]]
    row = [entry.value for entry in sheet[2]]
    assert row[headers.index("Prénom")] == "Jean"
    assert row[headers.index("Nom")] == "DUPONT"

    # Correction manuelle d'un champ absent de l'export (ex. la série) :
    # la valeur corrigée doit apparaître dans sa colonne.
    client.patch(
        f"/api/batches/{job['id']}/files/{file_id}/form",
        json={"page": 1, "values": {"serie": "042"}},
    )
    response = client.get(f"/api/batches/{job['id']}/export.xlsx")
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook["Données Extraites"]
    headers = [entry.value for entry in sheet[1]]
    row = [entry.value for entry in sheet[2]]
    assert row[headers.index("Série")] == "042"


def test_form_overrides_patch_and_excel(client: TestClient, tmp_path: Path):
    """Les corrections manuelles du formulaire doivent persister et être
    prises en compte dans l'export Excel."""
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]
    _wait_done(client, job["id"])

    items = client.get(f"/api/batches/{job['id']}/files").json()["items"]
    file_id = items[0]["id"]

    detail = client.get(f"/api/batches/{job['id']}/files/{file_id}").json()
    assert detail["pages"], "le fichier doit avoir une page analysée"
    assert detail["pages"][0]["form"] is not None

    # Pas de correction au départ.
    assert not detail.get("overrides")

    # Correction du champ "nom" sur la page 1.
    response = client.patch(
        f"/api/batches/{job['id']}/files/{file_id}/form",
        json={"page": 1, "values": {"nom": "DUPONT CORRIGE"}},
    )
    assert response.status_code == 200, response.text
    assert response.json()["overrides"]["nom"] == "DUPONT CORRIGE"

    detail = client.get(f"/api/batches/{job['id']}/files/{file_id}").json()
    assert detail["overrides"]["1"]["nom"] == "DUPONT CORRIGE"

    # Le détail refusé au client doit refléter la correction (statut valid).
    nom = next(
        field
        for field in detail["pages"][0]["form"]["fields"]
        if field["key"] == "nom"
    )
    assert nom["value"] == "DUPONT CORRIGE"
    assert nom["status"] == "valid"
    assert nom["edited"] is True

    # Une valeur vide efface la correction.
    response = client.patch(
        f"/api/batches/{job['id']}/files/{file_id}/form",
        json={"page": 1, "values": {"nom": ""}},
    )
    assert response.status_code == 200
    detail = client.get(f"/api/batches/{job['id']}/files/{file_id}").json()
    assert "1" not in detail.get("overrides", {})

    # Une seule page <= 1, page 99 introuvable.
    assert (
        client.patch(
            f"/api/batches/{job['id']}/files/{file_id}/form",
            json={"page": 99, "values": {"nom": "x"}},
        ).status_code
        == 404
    )


def test_cancel_and_delete_batch(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    job = _create_batch(client, [image])["job"]

    cancelled = client.post(f"/api/batches/{job['id']}/cancel")
    assert cancelled.status_code == 200
    assert cancelled.json()["cancelled"] is True

    deleted = client.delete(f"/api/batches/{job['id']}")
    assert deleted.status_code == 200
    assert deleted.json()["deleted"] == job["id"]

    assert client.get(f"/api/batches/{job['id']}").status_code == 404


def test_history_lists_jobs(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    first = _create_batch(client, [image], "Lot A")["job"]
    second = _create_batch(client, [image], "Lot B")["job"]
    jobs = client.get("/api/batches").json()["jobs"]
    ids = [entry["id"] for entry in jobs]
    assert first["id"] in ids and second["id"] in ids


# --------------------------------------------------------------------------- #
# Validation des entrées
# --------------------------------------------------------------------------- #
def test_create_batch_rejects_empty(client: TestClient):
    response = client.post("/api/batches", files=[])
    assert response.status_code in (400, 422)


def test_create_batch_rejects_all_invalid_files(client: TestClient, tmp_path: Path):
    bad = tmp_path / "note.txt"
    bad.write_text("pas une image")
    response = client.post(
        "/api/batches",
        files=[("files", ("note.txt", bad.read_bytes(), "text/plain"))],
    )
    assert response.status_code == 400


def test_create_batch_isolates_invalid_files(client: TestClient, tmp_path: Path):
    image = _sample_image(tmp_path)
    bad = tmp_path / "note.txt"
    bad.write_text("pas une image")
    response = client.post(
        "/api/batches",
        files=[
            ("files", ("scan.png", image.read_bytes(), "image/png")),
            ("files", ("note.txt", bad.read_bytes(), "text/plain")),
        ],
    )
    assert response.status_code == 201
    payload = response.json()
    assert len(payload["rejected"]) == 1
    assert payload["job"]["counts"]["total"] == 1


def test_unknown_job_returns_404(client: TestClient):
    assert client.get("/api/batches/nope").status_code == 404
    assert client.get("/api/batches/nope/files").status_code == 404


# --------------------------------------------------------------------------- #
# Lecture par zones (profil SCRIPTVAULT_ROI)
# --------------------------------------------------------------------------- #
def test_batch_roi_profile_passes_rois_to_engine(tmp_path: Path):
    """Avec ``roi_profile`` configuré, le lot transmet ``rois`` + le
    ``scan_barcode=False`` au moteur (lecture par champs du formulaire)."""
    captured: dict = {}

    class RoiEngine:
        cpu_threads = 4
        is_ready = True

        def predict_array(self, image, *, preprocess=True, rois=None, scan_barcode=None):
            captured["rois"] = rois
            captured["scan_barcode"] = scan_barcode
            return [
                {
                    "label": label,
                    "text": "VALEUR",
                    "confidence": 0.8,
                    "box": [[0, 0], [50, 0], [50, 10], [0, 10]],
                }
                for label in (rois or {})
            ]

        def close(self):
            pass

    settings = Settings(
        preload=False,
        max_concurrency=1,
        timeout_ms=5000,
        max_file_mb=5,
        preprocess=True,
        storage_root=str(tmp_path / "storage"),
        roi_profile={"nom": (0.02, 0.09, 0.6, 0.14), "cin": (0.02, 0.24, 0.5, 0.29)},
    )
    app = create_app(settings=settings, engine_factory=lambda: RoiEngine())
    image = _sample_image(tmp_path)
    with TestClient(app) as client:
        job = _create_batch(client, [image])["job"]
        _wait_done(client, job["id"])
        detail = client.get(f"/api/batches/{job['id']}/files")
        assert detail.status_code == 200
    assert captured["rois"] == settings.roi_profile
    assert captured["scan_barcode"] is False
    assert client.post("/api/batches/nope/cancel").status_code == 404
    assert client.delete("/api/batches/nope").status_code == 404
    assert client.get("/api/batches/nope/export.xlsx").status_code == 404
